# -*- coding: utf-8 -*-
import paramiko, re, os
from openpyxl import Workbook
# solve UnicodeDecodeError
import sys
reload(sys)
sys.setdefaultencoding('utf8')

class ssh_connect():
    def __init__(self,host,username,password):
        self.name = username
        self.host = host
        self.pwd = password
        self.ssh = paramiko.SSHClient()
        #  加上这句话不用担心选yes的问题，会自动选上
        # （用ssh连接远程主机时，第一次连接时会提示是否继续进行远程连接，选择yes）
        self.ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        self.ssh.connect(self.host, 22, self.name, self.pwd, timeout=10)  # 连接远程主机，SSH端口号为22
    # execute command
    def ssh_command(self, cmd):
        stdin, stdout, stderr = self.ssh.exec_command(cmd)
        return stdout.readlines()

    # write to result file
    def write_to_file(self, file, *info_list):
        with open(file, 'w')  as f:
            f.writelines(str(line) + "\n" for line in info_list)

    def line_prepender(self, filename, line):
        with open(filename, 'r+') as f:
            content = f.read()
            f.seek(0, 0)
            f.write(line.rstrip('\r\n') + '\n' + content)

    # get cmd result
    # k for every item, v for the relevant command
    def get_cmd_result(self, cmd_dict):
        result_cmd_dict = {}
        for k, v in cmd_dict.iteritems():
            output = self.ssh_command(v)
            # print output
            if len(output) < 2:
                # remove replace('\n', '') when the number of vm machine is less than 2 to ensure that output is a list
                if k == "cmd3":
                    result_cmd_dict[k] = output
                else:
                    result_cmd_dict[k] = output[0].replace('\n', '')
            else:
                result_cmd_dict[k] = output
        return result_cmd_dict

    def get_host_info(self):
        host_info_list = []
        # k for every item, v for the relevant command
        # cmd3 indicate the vm machine name
        cmds = {
            "cmd1" : "hostname |awk -F '[.]' '{print $1}'",
            "cmd2" : "esxcli hardware cpu list |grep CPU: |wc -l",
            "cmd3": "vim-cmd vmsvc/getallvms |awk 'NR>1{print $1,$2}' |sort -nr",
            "cmd4" : "esxcli hardware memory get |awk 'NR==1{printf \"%d\\n\",$(NF-1)/1024/1024/1024+1}'",
            "cmd5": "df  |awk 'NR>1{sum+=$2}END{printf \"%0.2f\",sum/1024/1024/1024/1024}'",
            "cmd6" : "esxcfg-info |grep 'Serial Number' |awk -F '[.]' '$1~/Serial/{print $NF}'",
            "cmd7": "esxcli hardware platform get |grep 'Product Name'|awk -F '[:]' '{print gensub(/ /,\"\",g,$NF)}'",
            "cmd8": "esxcli system version get |grep Version |awk '{print $NF}'"
        }
        result_dict = self.get_cmd_result(cmds)
        host_info = '{0} {1} {2} {3} {4} {5} {6} {7}'.format(result_dict['cmd1'], self.host,
                                                   result_dict['cmd2'], result_dict['cmd4'], result_dict['cmd5'],
                                                   result_dict['cmd6'], result_dict['cmd7'], result_dict['cmd8'])
        host_info_list.append(host_info)
        return host_info_list

    # get host info
    def get_vm_info(self):
        # k for every item, v for the relevant command
        # cmd3 indicate the vm machine name
        cmds = {
            "cmd1" : "hostname |awk -F '[.]' '{print $1}'",
            "cmd3" : "vim-cmd vmsvc/getallvms |awk 'NR>1 && NF>1 && $1~/[0-9]{1,3}/{print $1,$2}' |sort -nr",
        }
        result_dict = self.get_cmd_result(cmds)
        print result_dict
        host_info_list = []
        # get vm info
        for i in result_dict['cmd3']:
            vm_vmid = i.replace('\n', '').split(' ')[0]
            vm_hostname = i.replace('\n', '').split(' ')[1]
            cmd_get_vm_info = "vim-cmd vmsvc/get.guest {0}" .format(vm_vmid)
            cmd_power_status = "vim-cmd vmsvc/get.summary {0} |grep powerState |awk -F '\"' '{{print $(NF-1)}}'" .format(vm_vmid)
            cmd_tool_status = "vim-cmd vmsvc/get.guest {0} |grep toolsStatus |awk -F '\"' '{{print $(NF-1)}}'" .format(vm_vmid)
            cmd_guest_status = "vim-cmd vmsvc/get.guest {0} |grep guestState |awk -F '\"' '{{print $(NF-1)}}'" .format(vm_vmid)
            # cmd of cpu_count,mem,disk size
            result_dict_vm = {}
            cmds_vm = {
                "cpu" : "vim-cmd vmsvc/get.summary {0} |grep numCpu |awk '{{split($NF,a,\",\");print a[1]}}'" .format(vm_vmid),
                "mem" : "vim-cmd vmsvc/get.summary {0} |grep memorySizeMB |awk '{{split($NF,a,\",\");printf\"%d\",a[1]/1024}}'" .format(vm_vmid),
                "disk" : "vim-cmd vmsvc/device.getdevices {0} |grep capacityInKB |awk -F '[, ]' '{{sum+=$(NF-2)}}END{{printf\"%d\",sum/1024/1024}}'" .format(vm_vmid)
            }
            vm_detailed_info = self.ssh_command(cmd_get_vm_info)
            vm_power_status = self.ssh_command(cmd_power_status)[0].replace('\n', '')
            vm_tool_status = self.ssh_command(cmd_tool_status)[0].replace('\n', '')
            vm_guest_status = self.ssh_command(cmd_guest_status)[0].replace('\n', '')
            print vm_vmid, vm_guest_status
            print self.ssh_command(cmd_guest_status)

            result_dict_vm = self.get_cmd_result(cmds_vm)
            #print vm_detailed_info
            ip_list = []
            if vm_power_status == 'poweredOn' and vm_guest_status == 'running':
                for i, line in enumerate(vm_detailed_info):
                    if 'hostName' in line and 'ipAddress' in vm_detailed_info[i+1]:
                        # lan_ip = vm_detailed_info[i+1].split('"')[-2]
                        lan_ip = re.split('["<>]', vm_detailed_info[i+1])[-2]
                    if 'ipAddress' in line and re.search(r'[0-9]{1,3}\.', line) and 'prefixLength' in vm_detailed_info[i+1]:
                        if lan_ip not in line:
                            other_ip = line.split('"')[-2]
                            ip_list.append(other_ip)
                all_other_ip = "," .join(ip_list)
                if not re.search(r'[0-9]{1,3}\.', all_other_ip):
                    all_other_ip = 'unset'
                host_info = '{0} {1} {2} {3} {4} {5} {6} {7} {8} {9}'.format(result_dict['cmd1'], self.host,
                                 vm_vmid, vm_hostname, vm_power_status, lan_ip, result_dict_vm['cpu'],
                                 result_dict_vm['mem'], result_dict_vm['disk'], all_other_ip)
            else:
                lan_ip = 'unset'
                all_other_ip = 'unset'
                host_info = '{0} {1} {2} {3} {4} {5} {6} {7} {8} {9}'.format(result_dict['cmd1'], self.host,
                                 vm_vmid, vm_hostname, vm_power_status, lan_ip, result_dict_vm['cpu'],
                                 result_dict_vm['mem'], result_dict_vm['disk'], all_other_ip)
            host_info_list.append(host_info)
        # write the vm device info to file named vm_info.txt
        first_line = '宿主机名称 宿主机IP 虚拟机vmid 主机名 电源状态 内网IP CPU核数 内存(G)  硬盘(G) 其他IP'
        self.write_to_file('vm_info.txt', *host_info_list)
        self.line_prepender('vm_info.txt', first_line)
        #print host_info_list
        return host_info_list

class Main(object):
    def __init__(self, ip_list_file, username, password):
        self.username = username
        self.password = password
        self.ip_list_file = ip_list_file
    # force the number in host_device_info to int or float
    def convert_list_format(self, info_list):
        info_new = []
        for x in info_list:
            if re.search(r'(^[0-9]{1,5}$)', x):
                x = int(x)
            elif re.search(r'(^[0-9]{1,2}\.[0-9]{2}$)', x):
                x = float(x)
            else:
                x = x
            info_new.append(x)
        return info_new
    # put all host and vm info to list named all_info_list
    def all_host_and_vm_info(self):
        # username = 'root'
        # password = 'xxxxxxx'
        # ip_list_file = 'ip.txt'
        with open(ip_list_file) as f:
            hosts_info_list = []
            vms_info_list = []
            all_info_list = []
            for host in f:
                try:
                    host = host.strip()
                    ssh = ssh_connect(host, username, password)
                    vm_info_list = ssh.get_vm_info()
                    host_info_list = ssh.get_host_info()
                    hosts_info_list.append(host_info_list)
                    vms_info_list.append(vm_info_list)
                except Exception, msg:
                    print '{0} bad msg, cannot connect to {1}' .format(msg, host)
            print hosts_info_list
            print vms_info_list
            all_info_list.append(hosts_info_list)
            all_info_list.append(vms_info_list)
            return all_info_list

    def print_host_lists_excel(self):
        all_info_list = self.all_host_and_vm_info()
        # write the info to excel
        wb = Workbook(write_only=True)
        ws = []
        host_tag_lists = ['宿主机', '虚拟机']
        # create sheet
        for i in range(len(host_tag_lists)):
            ws.append(wb.create_sheet(title=host_tag_lists[i].decode()))  # utf8->unicode
        # insert sheet header
        ws[0].append(['序号', '主机名', 'IP', 'CPU核数', '内存(G)', '硬盘(T)', '序列号', '型号', 'ESXI版本号'])
        ws[1].append(['序号', '宿主机名称', '宿主机IP', '虚拟机vmid', '主机名', '电源状态', '内网IP', 'CPU核数', '内存(G)', '硬盘(G)', '其他IP'])
        # insert host and vm info
        for i in range(len(host_tag_lists)):
            count = 1
            for info_list in all_info_list[i]:
                # traverse the hosts_info_list and vms_info_list
                for info in info_list:
                    print 'info: {0}' .format(info)
                    info = info.split(' ')
                    info_new = self.convert_list_format(info)
                    info_new.insert(0, count)
                    print 'info_list_with_seq_num: {0}' .format(info_new)
                    ws[i].append(info_new)
                    # ws[0].append([count, info[0], info[1], int(info[2]), int(info[3]), float(info[4]), info[5], info[6], info[7]])
                    count += 1
        # define the filename and save it to local disk
        save_path = 'host_list'
        for i in range(len(host_tag_lists)):
            save_path += ('-' + host_tag_lists[i].decode())
        save_path += '.xlsx'
        wb.save(save_path)


if __name__ == '__main__':
    username = 'root'
    password = 'xxxxxxx'
    ip_list_file = 'ip.txt'
    main = Main(ip_list_file, username, password)
    #main.all_host_and_vm_info()
    main.print_host_lists_excel()
